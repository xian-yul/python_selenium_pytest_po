import re
import time
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
import pyautogui
from selenium.webdriver.common import keys
import random

from datetime import datetime
# coding=utf-8

import os
import subprocess
import traceback
import logging

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait


def input_clear_text(driver, method, loc, text):
    loc_object = driver.find_element(method, loc)
    loc_object.clear()
    loc_object.send_keys(keys.Keys.CONTROL, "a")
    for i in range(10):
        loc_object.send_keys(keys.Keys.BACKSPACE)
    loc_object.send_keys(text)


def random_number(num):
    str = ""
    for i in range(num):
        ch = chr(random.randrange(ord('0'), ord('9') + 1))
        str += ch
    return str


def execute_script(driver, size):
    js = "var q=document.documentElement.scrollTop=" + size
    driver.execute_script(js)


# 比较俩个时间的差异
def time_lag(time, times):
    d1 = datetime.strptime(time, '%Y-%m-%d %H:%M:%S')
    d2 = datetime.strptime(times, '%Y-%m-%d %H:%M:%S')
    delta = d1 - d2
    return delta


# 原料牌号返回
def goods_mark(num):
    # 选择 原料牌号
    yl_mark = "div.ant-modal-body div:nth-child(" + str(num) + ")"
    return yl_mark


def goods_deliver_area(area):
    deliver_area = "xpath==//*[contains(text()," + area + ")]"
    return deliver_area


def remove_chinese(text):
    pattern = '[\u4e00-\u9fff]'  # 匹配非中文字符
    return re.sub(pattern, '', text)


def unicode_name():
    val = random.randint(0x4e00, 0x9fbf)
    return chr(val)


def gbk2312_name():
    head = random.randint(0xb0, 0xf7)
    body = random.randint(0xa1, 0xfe)
    val = f'{head:x} {body:x}'
    str = bytes.fromhex(val).decode('gb2312')
    return str


def look_for_text(loc):
    xpath = "//*[contains(text()," + loc + ")]"
    return xpath


TESSERACT = 'tesseract'  # 调用的本地命令名称
TEMP_IMAGE_NAME = "temp.bmp"  # 转换后的临时文件
TEMP_RESULT_NAME = "temp"  # 保存识别文字临时文件
CLEANUP_TEMP_FLAG = True  # 清理临时文件的标识
INCOMPATIBLE = True  # 兼容性标识


def image_to_scratch(image, TEMP_IMAGE_NAME):
    # 将图片处理为兼容格式
    image.save(TEMP_IMAGE_NAME, dpi=(200, 200))


def retrieve_text(TEMP_RESULT_NAME):
    # 读取识别内容
    inf = open(TEMP_RESULT_NAME + '.txt', 'r')
    text = inf.read()
    inf.close()
    return text


def perform_cleanup(TEMP_IMAGE_NAME, TEMP_RESULT_NAME):
    # 清理临时文件
    for name in (TEMP_IMAGE_NAME, TEMP_RESULT_NAME + '.txt', "tesseract.log"):
        try:
            os.remove(name)
        except OSError:
            pass


def call_tesseract(image, result, lang):
    # 调用tesseract.exe，将识读结果写入output_filename中
    args = [TESSERACT, image, result, '-l', lang]
    proc = subprocess.Popen(args)
    retcode = proc.communicate()


def get_random_chinese_characters(num):
    characters = []
    for _ in range(num):
        # 根据Unicode编码范围筛选汉字
        char = chr(random.randint(0x4E00, 0x9FBF))
        characters.append(char)
    return characters


def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 获取表单
    max_raw = sheet.max_row  # 获取最大行数
    print(max_raw)
    case_list = []  # 创建空列表 存放测试用例
    for i in range(2, max_raw + 1):
        dict1 = dict(
            name=sheet.cell(row=i, column=1).value,
        )
        case_list.append(dict1)
    return case_list


def input_clear_text(loc_object, text):
    loc_object.clear()
    loc_object.send_keys(keys.Keys.CONTROL, "a")
    for i in range(10):
        loc_object.send_keys(keys.Keys.BACKSPACE)
    loc_object.send_keys(text)


def signing_contract(driver, time):
    iframe = return_element(driver, "// iframe[@id='fddContractSigning']")
    driver.switch_to.frame(iframe)
    print('定位签署框')
    ec_element(driver, "// div[@class='sign-btn-box-con']")
    # driver.find_element(By.XPATH, "// div[@class='sign-btn-box-con']").click()
    time.sleep(2)
    print('点击签署按钮')
    input_clear_text(driver.find_element(By.XPATH, "//input[@id='smsval']"), 999999)
    ec_element(driver, "// a[@class='commit sign-btn fontbg-config-style en_sms_confirm']")
    time.sleep(5)
    driver.switch_to.default_content()
    time.sleep(5)
    ec_element(driver, "// button[@class='ant-modal-close']")
    time.sleep(3)


def ec_all_elements(driver, loc, subscript):
    loc_click = WebDriverWait(driver, 100).until(
        EC.visibility_of_all_elements_located((By.XPATH, loc))
    )
    loc_click[subscript].click()


def ec_element(driver, loc):
    loc_click = WebDriverWait(driver, 100).until(
        EC.element_to_be_clickable((By.XPATH, loc))
    )
    loc_click.click()


def return_element(driver, loc):
    loc = WebDriverWait(driver, 100).until(
        EC.element_to_be_clickable((By.XPATH, loc))
    )
    return loc
